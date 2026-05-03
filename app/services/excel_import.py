import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
import sqlite3

from app.services.membership_service import sanitize_input

def sanitize_cell_value(value) -> Optional[str]:
    if value is None:
        return None
    cleaned = str(value).strip() if str(value).strip() else None
    if cleaned:
        return sanitize_input(cleaned)
    return None

def convert_bool_value(value) -> bool:
    if isinstance(value, str):
        return value.lower() in ['yes', 'true', '1', 'y']
    return bool(value)

def find_existing_member(conn, row_data: Dict[str, Any]) -> Optional[int]:
    cursor = conn.cursor()
    
    # Try membership_number first (highest priority)
    if row_data.get('membership_number'):
        cursor.execute("SELECT id FROM members WHERE membership_number = ?", 
                       (row_data['membership_number'],))
        result = cursor.fetchone()
        if result:
            return result['id']
    
    # Try email second
    if row_data.get('email') and row_data['email'] != 'N/A':
        cursor.execute("SELECT id FROM members WHERE email = ?", (row_data['email'],))
        result = cursor.fetchone()
        if result:
            return result['id']
    
    # Try name match (combined first_name + last_name)
    if row_data.get('name'):
        cursor.execute("SELECT id FROM members WHERE name = ?", (row_data['name'],))
        results = cursor.fetchall()
        if len(results) == 1:
            return results[0]['id']
    
    # Try first_name + last_name combo
    if row_data.get('first_name') and row_data.get('last_name'):
        cursor.execute("SELECT id FROM members WHERE first_name = ? AND last_name = ?",
                          (row_data['first_name'], row_data['last_name']))
        results = cursor.fetchall()
        if len(results) == 1:
            return results[0]['id']
    
    return None

def merge_member_data(conn, member_id: int, row_data: Dict[str, Any]):
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM members WHERE id = ?", (member_id,))
    existing = dict(cursor.fetchone())
    
    update_fields = {}
    
    # Handle membership_amount_used separately
    if 'membership_amount_used' in row_data and row_data['membership_amount_used'] is not None:
        if not existing.get('amount_used') and row_data['membership_amount_used']:
            update_fields['amount_used'] = float(row_data['membership_amount_used'])
    
    # Handle N/A email
    email = row_data.get('email')
    if email == 'N/A':
        email = None
    
    for field in ['membership_number', 'name', 'first_name', 'last_name', 'email', 
                  'membership_type', 'includes_cart', 'includes_range']:
        value = email if field == 'email' else row_data.get(field)
        if field in row_data or (field == 'email' and email):
            if not existing.get(field) and value:
                update_fields[field] = value
    
    if update_fields:
        set_clause = ", ".join(f"{k} = ?" for k in update_fields)
        values = list(update_fields.values()) + [datetime.now(timezone.utc).isoformat(), member_id]
        cursor.execute(f"UPDATE members SET {set_clause}, last_updated = ? WHERE id = ?", values)
        conn.commit()
    
    return update_fields

def insert_new_member(conn, row_data: Dict[str, Any], source_sheet: str):
    cursor = conn.cursor()
    
    name = row_data.get('name')
    if not name and row_data.get('first_name') and row_data.get('last_name'):
        name = f"{row_data['first_name']} {row_data['last_name']}"
    
    # Handle "N/A" email values
    email = row_data.get('email')
    if email == 'N/A':
        email = None
    
    # Get amount used from column "membership_amount_used"
    amount_used = row_data.get('membership_amount_used', 0)
    if amount_used is None:
        amount_used = 0
    
    cursor.execute("""
    INSERT INTO members (membership_number, name, first_name, last_name, email, 
                        membership_type, amount_used, includes_cart, includes_range, 
                        source_sheet, last_updated)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        row_data.get('membership_number'),
        name,
        row_data.get('first_name'),
        row_data.get('last_name'),
        email,
        row_data.get('membership_type'),
        float(amount_used) if amount_used else 0,
        convert_bool_value(row_data.get('includes_cart', False)),
        convert_bool_value(row_data.get('includes_range', False)),
        source_sheet,
        datetime.now(timezone.utc).isoformat()
    ))
    conn.commit()
    return cursor.lastrowid

def import_excel_file(conn, filepath: str, progress_callback=None) -> Dict[str, Any]:
    wb = load_workbook(filepath, data_only=True)
    results = {"sheets_processed": 0, "members_updated": 0, "members_inserted": 0, "errors": []}
    
    for sheet_name in wb.sheetnames:
        if 'summary' in sheet_name.lower():
            continue
        
        ws = wb[sheet_name]
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Filter out None headers and create a mapping of col_idx -> field_name
        col_mapping = {}
        for idx, h in enumerate(headers):
            if h is not None:
                field_name = str(h).lower().replace(' ', '_')
                col_mapping[idx] = field_name
        
        if not col_mapping:
            continue
        
        results["sheets_processed"] += 1
        row_count = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_count += 1
            if progress_callback:
                progress_callback(f"Processing {sheet_name}: row {row_idx}")
            
            row_data = {}
            for idx, cell_value in enumerate(row):
                if idx in col_mapping:
                    field_name = col_mapping[idx]
                    row_data[field_name] = sanitize_cell_value(cell_value)
            
            # Only process rows that have actual data (at least name or email)
            has_data = row_data.get('first_name') or row_data.get('last_name') or row_data.get('email')
            if not has_data:
                continue
            
            # Build full name from first/last if not already set
            if not row_data.get('name'):
                first = row_data.get('first_name', '')
                last = row_data.get('last_name', '')
                if first or last:
                    row_data['name'] = f"{first} {last}".strip()
            
            if not row_data.get('name') and not row_data.get('email'):
                continue
            
            if '_first_name' in row_data and '_last_name' in row_data:
                row_data['first_name'] = row_data.pop('_first_name')
                row_data['last_name'] = row_data.pop('_last_name')
            
            try:
                existing_id = find_existing_member(conn, row_data)
                if existing_id:
                    updated = merge_member_data(conn, existing_id, row_data)
                    if updated:
                        results["members_updated"] += 1
                else:
                    insert_new_member(conn, row_data, sheet_name)
                    results["members_inserted"] += 1
            except Exception as e:
                results["errors"].append(f"Row {row_idx} in {sheet_name}: {str(e)}")
        
        if progress_callback:
            progress_callback(f"Completed {sheet_name}: {row_count} rows")
    
    return results
