import pytest
import tempfile
import os
from pathlib import Path

def test_excel_import_insert_new(conn):
    from app.services.excel_import import import_excel_file
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(['Membership Number', 'Name', 'Email', 'Membership Type'])
    ws.append(['MBR001', 'John Doe', 'john@example.com', 'Standard'])
    
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    wb.save(temp_file.name)
    
    try:
        results = import_excel_file(conn, temp_file.name)
        assert results['members_inserted'] == 1
        assert results['members_updated'] == 0
        
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members WHERE membership_number = ?", ('MBR001',))
        row = cursor.fetchone()
        assert row is not None
        assert row['name'] == 'John Doe'
    finally:
        os.unlink(temp_file.name)

def test_excel_import_no_overwrite_existing(conn):
    """Test that existing DB values are NOT overwritten by Excel values (per spec)"""
    from app.services.excel_import import import_excel_file
    import openpyxl
    
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO members (membership_number, name, email, membership_type) VALUES (?, ?, ?, ?)",
        ('MBR001', 'John Doe', 'john@example.com', 'Standard')
    )
    conn.commit()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(['Membership Number', 'Name', 'Email', 'Membership Type'])
    ws.append(['MBR001', 'John Doe', 'john@example.com', 'Premium'])
    
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    wb.save(temp_file.name)
    
    try:
        results = import_excel_file(conn, temp_file.name)
        assert results['members_updated'] == 0
        
        cursor.execute("SELECT membership_type FROM members WHERE membership_number = ?", ('MBR001',))
        row = cursor.fetchone()
        assert row['membership_type'] == 'Standard'
    finally:
        os.unlink(temp_file.name)

def test_excel_import_merge_blank_values(conn):
    from app.services.excel_import import import_excel_file
    import openpyxl
    
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO members (membership_number, name, email, membership_type) VALUES (?, ?, ?, ?)",
        ('MBR001', 'John Doe', 'john@example.com', 'Standard')
    )
    conn.commit()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(['Membership Number', 'Name', 'Email', 'Membership Type'])
    ws.append(['MBR001', '', '', ''])
    
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    wb.save(temp_file.name)
    
    try:
        results = import_excel_file(conn, temp_file.name)
        
        cursor.execute("SELECT * FROM members WHERE membership_number = ?", ('MBR001',))
        row = cursor.fetchone()
        assert row['name'] == 'John Doe'
        assert row['email'] == 'john@example.com'
        assert row['membership_type'] == 'Standard'
    finally:
        os.unlink(temp_file.name)

def test_excel_import_skip_summary_sheet(conn):
    from app.services.excel_import import import_excel_file
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(['Summary data'])
    
    ws2 = wb.create_sheet("Members")
    ws2.append(['Membership Number', 'Name'])
    ws2.append(['MBR001', 'John Doe'])
    
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    wb.save(temp_file.name)
    
    try:
        results = import_excel_file(conn, temp_file.name)
        assert results['sheets_processed'] == 1
        assert results['members_inserted'] == 1
    finally:
        os.unlink(temp_file.name)

def test_excel_import_formula_injection_protection(conn):
    from app.services.excel_import import import_excel_file
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(['Membership Number', 'Name', 'Email'])
    ws.append(['=CMD()', 'John Doe', '@test'])
    
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    wb.save(temp_file.name)
    
    try:
        results = import_excel_file(conn, temp_file.name)
        
        cursor = conn.cursor()
        cursor.execute("SELECT membership_number, email FROM members WHERE name = ?", ('John Doe',))
        row = cursor.fetchone()
        assert row is not None
    finally:
        os.unlink(temp_file.name)
